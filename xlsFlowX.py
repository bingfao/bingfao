
#####################################################################
# use openpyxl to deal with excel file to export .h and .svh for DV
#####################################################################

# Author: binga.gao
# date: 2023-10-24
# change-desc:
#             change-date: 2023-10-24
#             1. change use xlrd to xlsx
#             2. show check error info in xlsx file with border color red
#             3. support virtual reg and  reg  group


import re
# from builtins import str
import os,sys
from openpyxl import load_workbook
from openpyxl.styles import colors, Border, Side, Font, Color


# import xlrd


char_arr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
busTypestr_arr = ("AHB", "AXI")
bitWidMask_arr = ('0x01', '0x03', '0x07', '0x0F', '0x1F', '0x3F', '0x7F', '0xFF', '0x01FF', '0x03FF', '0x07FF', '0x0FFF', '0x1FFF', '0x3FFF', '0x7FFF', '0xFFFF',
                  '0x01FFFF', '0x03FFFF', '0x07FFFF', '0x0FFFFF', '0x1FFFFF', '0x3FFFFF', '0x7FFFFF', '0xFFFFFF', '0x01FFFFFF', '0x03FFFFFF', '0x07FFFFFF', '0x0FFFFFFF', '0x1FFFFFFF', '0x3FFFFFFF', '0x7FFFFFFF', '0xFFFFFFFF',
                  '0x01FFFFFF', '0x03FFFFFF', '0x07FFFFFF', '0x0FFFFFFF', '0x1FFFFFFF', '0x3FFFFFFF', '0x7FFFFFFF', '0xFFFFFFFF', '0x01FFFFFF', '0x03FFFFFFFF', '0x07FFFFFFFF', '0x0FFFFFFFFF', '0x1FFFFFFFFF', '0x3FFFFFFFFF', '0x7FFFFFFFFF', '0xFFFFFFFFFF',
                  '0x01FFFFFFFF', '0x03FFFFFFFF', '0x07FFFFFFFF', '0x0FFFFFFFFF', '0x1FFFFFFFFF', '0x3FFFFFFFFF', '0x7FFFFFFFFF', '0xFFFFFFFFFF', '0x01FFFFFFFF', '0x03FFFFFFFFFF', '0x07FFFFFFFFFF', '0x0FFFFFFFFFFF', '0x1FFFFFFFFFFF', '0x3FFFFFFFFFFF', '0x7FFFFFFFFFFF', '0xFFFFFFFFFFFF')

fieldRWOp_arr = ('rw', 'ro', 'wo', 'w1', 'w1c', 'rc', 'rs', 'wrc', 'wrs', 'wc', 'ws', 'wsrc', 'wcrs', 'w1s', 'w1t', 'w0c',
                 'w0s', 'w0t', 'w1src', 'w1crs', 'w0src', 'w0crs', 'woc', 'wos', 'wo1')

# RW, RO, WO, W1, W1C, RC, RS, WRC, WRS, WC, WS, WSRC, WCRS, W1S, W1T, W0C,
#                  W0S, W0T, W1SRC, W1CRS, W0SRC, W0CRS, W0C, W0S, WO1

# uint_type_arr = ('uint8_t', 'uint16_t', 'uint32_t', 'uint64_t')


class St_Filed_info:
    def __init__(self, name, attr):
        self.end_bit = 31
        self.start_bit = 0
        self.attribute = attr
        self.defaultValue = 0
        self.field_name = name
        self.field_comments = ''
        self.field_enumstr = ''
        self.field_constr = ''
        self.hdl_path = ''
        self.bRandom_Enable = False
        self.xls_row = 0

    def field_info_str(self):
        out_str = f'fieldname: {self.field_name}, end_bit: {self.end_bit}, start_bit: {self.start_bit}, attribute: {self.attribute} \n , defaultValue: {hex(self.defaultValue)}, comments: {self.field_comments}, enum: {self.field_enumstr}'
        return out_str


class St_Reg_info:
    def __init__(self, name):
        self.field_list = []
        self.reg_name = name
        self.offset = 0
        self.desc = ''          # 寄存器的描述
        self.bVirtual = False
        self.bGroup_start = 0   # 是否是多组的起始
        self.bGroup_stop = 0    # 是否是多组的结束
        self.group_dim = 0      # 有几组
        self.group_size = 0     # reg组的size
        self.group_name = ''
        self.group_index = -1

    def field_count(self):
        return len(self.field_list)

    def is_fieldInReg(self, fd_name):
        bIn = fd_name == self.reg_name
        if not bIn:
            for fd in self.field_list:
                if fd.field_name == fd_name:
                    bIn = True
                    break
        return bIn

    def add_field(self, fd):
        self.field_list.append(fd)

    def getCHeaderString(self):
        return ""

    def reg_info_str(self):
        group_info = ''
        if self.group_index >= 0:
            group_info = f'group: {self.group_name}, group_index: {self.group_index}, '
        out_str = f'regname: {self.reg_name}, offset_addr: {hex(self.offset)}, {group_info} virtual: {self.bVirtual}'

        # for f in self.field_list:
        #     out_str += "\n"+f.field_info_str()
        return out_str


class St_Module_info:
    def __init__(self, name):
        self.module_name = name
        self.bus_baseAddr = 0
        self.addr_width = 32
        self.data_width = 32
        self.bus_type = 0
        self.hdl_path = 'NULL'
        self.reg_list = []

    def reg_count(self):
        return len(self.reg_list)

    def getAllFieldCount(self):
        nCount = 0
        for r in self.reg_list:
            nCount += r.field_count()
        return nCount

    def appendRegInfo(self, reginfo):
        self.reg_list.append(reginfo)

    def getCHeaderString(self):
        return ""

    def getCSourceString(self):
        return ""

    def module_info_str(self):
        out_str = f'moduleName: {self.module_name}, bus_type: {self.bus_type}, bus_addr: {hex(self.bus_baseAddr)}'
        for r in self.reg_list:
            out_str += "\n"+r.reg_info_str()
        return out_str


def markCell_InvalidFunc2(ws, row, col, clr='ff0000'):
    double = Side(border_style="double", color=clr)
    border = Border(left=double,
                    right=double,
                    top=double,
                    bottom=double)
    cell = ws.cell(row, col)
    cell.border = border
    cell.font = Font(color="FF0000")


def markCell_InvalidFunc(ws, cellstr, clr='ff0000'):
    double = Side(border_style="double", color=clr)
    border = Border(left=double,
                    right=double,
                    top=double,
                    bottom=double)
    cell = ws[cellstr]
    if not isinstance(cell, tuple):
        cell.border = border
        cell.font = Font(color="FF0000")


def isUnallowedVarName(strVal):
    # strVal = strVal.strip()
    pattern = '^[a-zA-Z_][a-zA-Z0-9_]*$'
    matchObject = re.search(pattern, strVal)
    # if matchObject is None:
    #     print('%s is not Id' % id)
    # else:
    #     print('%s is Id' % id)
    return (matchObject is None)


def isHexString(strVal, b0xStart=True):
    strUpper = strVal.upper()
    brt = True
    if b0xStart:
        brt = strUpper.startswith('0X')
    hexstr = '0123456789ABCDEF'
    if brt:
        strhex = strUpper[2:]
        for c in strhex:
            if c not in hexstr:
                brt = False
                break
    return brt

def get_output_c_dir():
    if sys.platform == 'linux':
        prj_root = os.getenv("PRJ_ROOT")
        result_dir    = os.path.join(prj_root,'dv/tb/reg_model/c')
        return result_dir

def get_output_dut_cfg_dir():
    if sys.platform == 'linux':
        prj_root = os.getenv("PRJ_ROOT")
        result_dir    = os.path.join(prj_root,'dv/tb/reg_model/sv')
        return result_dir

def get_output_ral_dir():
    if sys.platform == 'linux':
        prj_root = os.getenv("PRJ_ROOT")
        result_dir    = os.path.join(prj_root,'dv/tb/reg_model/ral')
        return result_dir

def checkModuleSheetVale(ws):  # 传入worksheet
    # print("Start Check Sheet Values.")
    modName = ws['B1'].value
    baseAddr0 = ws['D1'].value
    baseAddr1 = ws['D2'].value
    data_width = ws['B2'].value
    # addr_with = ws['D2'].value
    ahb_hdl_path = ws['F1'].value
    axi_hdl_path = ws['F2'].value
    bCheckPass = True
    bExcelBasePass = True
    st_module_list = []
    if modName is None:
        print("ModuleName must be filled.")
        markCell_InvalidFunc(ws, 'B1')
        bExcelBasePass = False
    elif isinstance(modName, str):
        if isUnallowedVarName(modName):
            print("ModuleName only can include letter,number,and _ in middle.")
            markCell_InvalidFunc(ws, 'B1')
            bExcelBasePass = False
    if baseAddr0 is None and baseAddr1 is None:
        print("baseAddr must be filled.")
        markCell_InvalidFunc(ws, 'F1')
        markCell_InvalidFunc(ws, 'H2')
        bExcelBasePass = False

    if data_width is None:
        print("daa_width must be filled.")
        markCell_InvalidFunc(ws, 'B2')
        bExcelBasePass = False
    # if addr_with is None:
    #     print("addr_witdh must be filled.")
    #     markCell_InvalidFunc(ws, 'D2')
    #     bExcelBasePass = False

    if bExcelBasePass:
        if baseAddr0 is not None:
            # print(baseAddr0)
            ahb_addr_lst = baseAddr0.splitlines()
            ahb_hdl_path_lst = []
            if ahb_hdl_path and ahb_hdl_path != 'NULL':
                ahb_hdl_path_lst = ahb_hdl_path.splitlines()
                if len(ahb_hdl_path_lst) != len(ahb_addr_lst):
                    print("hdl_path must be same size as addr.")
                    markCell_InvalidFunc(ws, 'F1')
                    bExcelBasePass = False
            i = 0
            for ahb in ahb_addr_lst:
                ahb_module = St_Module_info(modName)
                ahb_module.bus_baseAddr = int(ahb, 16)
                ahb_module.data_width = data_width
                if len(ahb_hdl_path_lst) > i:
                    ahb_module.hdl_path = ahb_hdl_path_lst[i]
                st_module_list.append(ahb_module)
                i += 1
        if baseAddr1 is not None:
            # print(baseAddr1)
            axi_addr_lst = baseAddr1.splitlines()
            axi_hdl_path_lst = []
            if axi_hdl_path and axi_hdl_path != 'NULL':
                axi_hdl_path_lst = axi_hdl_path.splitlines()
                if len(axi_addr_lst) != len(axi_hdl_path_lst):
                    print("hdl_path must be same size as addr.")
                    markCell_InvalidFunc(ws, 'F2')
                    bExcelBasePass = False
            i = 0
            for axi in axi_addr_lst:
                axi_module = St_Module_info(modName)
                axi_module.bus_baseAddr = int(axi, 16)
                axi_module.bus_type = 1
                axi_module.data_width = data_width
                if len(axi_hdl_path_lst) > i:
                    axi_module.hdl_path = axi_hdl_path_lst[i]
                st_module_list.append(axi_module)
                i += 1
    if bExcelBasePass == False:
        bCheckPass = False

    nRows = ws.max_row
    # nCols = ws.ncols
    regNameList = []

    bRegCheckPass = True
    # 表格固定从起始行  6开始
    i = 6
    laststartBit = 0
    bNeedNewReg = True
    # b_NewRegName = False
    reg_info = St_Reg_info('error')
    # while i < nRows:
    #     bSkip = False
    group_index = -1
    group_name = ''
    for row in ws.iter_rows(min_row=6, max_col=19, max_row=nRows, values_only=True):
        if all(cell is None for cell in row):
            bSkip = True
        else:
            # regName = ws.cell(i, 1).value
            # regOffset = ws.cell(i, 8).value
            regName = row[0]
            if bNeedNewReg:
                if regName is None:
                    print("Cell[A"+str(i)+"] regName must be filled.")
                    markCell_InvalidFunc(ws, f'A{i}')
                    bRegCheckPass = False

            if isinstance(regName, str):
                regName = regName.strip()
                if len(regName) == 0:
                    print("Cell[A"+str(i)+"] regName is empty string, Not allowed.")
                    markCell_InvalidFunc(ws, f'A{i}')
                    bRegCheckPass = False
                elif isUnallowedVarName(regName):
                    print(f'A{i} '+f'regName \"{regName}\" only can include letter,number,and _ in middle.')
                    markCell_InvalidFunc(ws, f'A{i}')
                    bRegCheckPass = False
                else:
                    regName = regName.upper()
                    # b_NewRegName = True
                    laststartBit = 0xFF
                    # virtual_val = ws.cell(i, 2).value
                    # group_start = ws.cell(i, 3).value
                    # group_stop = ws.cell(i, 4).value
                    # group_dim = ws.cell(i, 5).value
                    # group_size = ws.cell(i, 6).value

                    virtual_val = row[1]
                    group_start = row[2]
                    group_stop = row[3]
                    group_dim = row[4]
                    group_size = row[5]
                    reg_desc = row[6]
                    regOffset = row[7]
                    if not (regName in regNameList):
                        regNameList.append(regName)

                        reg_info = St_Reg_info(regName)
                        reg_info.bVirtual = (virtual_val == 1)
                        reg_info.bGroup_start = (group_start == 1)
                        reg_info.bGroup_stop = (group_stop == 1)
                        if reg_info.bGroup_start:
                            if not (group_dim and group_size):
                                print("Cell[H"+str(i)+"] dim and dim_size must be filled.")
                                if not group_dim:
                                    markCell_InvalidFunc(ws, f'E{i}')
                                if not group_size:
                                    markCell_InvalidFunc(ws, f'F{i}')
                                bRegCheckPass = False
                            reg_info.group_dim = group_dim
                            reg_info.group_size = group_size
                            reg_info.group_index = 0
                            group_name = f'st_group_{regName}'
                            if not reg_info.bGroup_stop:
                                group_index = 0
                            else:
                                reg_info.group_name = group_name
                        elif reg_info.bGroup_stop:
                            reg_info.group_index = group_index+1
                            group_name += f'__{regName}'
                            for module in st_module_list:
                                for i in range(0 - reg_info.group_index, 0):
                                    reg = module.reg_list[i]
                                    reg.group_name = group_name
                            reg_info.group_name = group_name
                            group_name = ''
                            group_index = -1
                        elif group_index > 0:
                            group_index += 1
                            reg_info.group_index = group_index
                        reg_info.desc = reg_desc
                        if not reg_info.bVirtual:
                            if regOffset is None:
                                print("Cell[H"+str(i)+"] offset Addr must be filled.")
                                markCell_InvalidFunc(ws, f'H{i}')
                                bRegCheckPass = False
                            if isinstance(regOffset, str):
                                # 待增加offset 值越来越大的规则判断
                                # print(regOffset)
                                if not isHexString(regOffset):
                                    print("Cell[F"+str(i)+"] offset Addr must be 0xFFFFFFF like hex string.")
                                    markCell_InvalidFunc(ws, f'H{i}')
                                    bRegCheckPass = False
                                else:
                                    reg_info.offset = int(regOffset, 16)
                                    if st_module_list:
                                        module = st_module_list[-1]
                                        if module.reg_count():
                                            lastOffset = module.reg_list[-1].offset
                                            if lastOffset >= reg_info.offset:
                                                print("Cell[H"+str(i)+"] offset Addr must > last reg offset.")
                                                markCell_InvalidFunc(ws, f'H{i}')
                                                bCheckPass = False
                                        for module in st_module_list:
                                            module.appendRegInfo(reg_info)
                        else:
                            for module in st_module_list:
                                module.appendRegInfo(reg_info)
                    else:
                        print("Cell[A"+str(i)+"] regName repeated, Not allowed.")
                        markCell_InvalidFunc(ws, f'A{i}')
                        bRegCheckPass = False

            # 处理每一行的 Field Info  这里需要重新考虑下
            bFiled_info_Pass = True

            if not bRegCheckPass:
                #     reg_info = st_module_list[-1].reg_list[-1]
                # else:
                bCheckPass = False

            # if not reg_info.bVirtual:
                # field_name = ws.cell(i, 10).value
                # endBit = ws.cell(i, 11).value
                # startBit = ws.cell(i, 12).value
                # field_attr = ws.cell(i, 13).value
            field_name = row[9]
            endBit = row[10]
            startBit = row[11]
            field_attr = row[12]

            if field_name is None:
                print("Cell[J"+str(i)+"] must be filled.")
                markCell_InvalidFunc(ws, f'J{i}')
                bFiled_info_Pass = False
            else:
                field_name = field_name.strip()
                field_name = field_name.upper()
                if isUnallowedVarName(field_name):
                    print(f'J{i} '+f'field_name \"{field_name}\"only can include letter,number,and _ in middle.')
                    markCell_InvalidFunc(ws, f'J{i}')
                    bFiled_info_Pass = False
            if endBit is None:
                print("Cell[K"+str(i)+"] must be filled.")
                markCell_InvalidFunc(ws, f'k{i}')
                bFiled_info_Pass = False
            if startBit is None:
                print("Cell[L"+str(i)+"] must be filled.")
                markCell_InvalidFunc(ws, f'L{i}')
                bFiled_info_Pass = False
            if field_attr is None:
                print("Cell[M"+str(i)+"] must be filled.")
                markCell_InvalidFunc(ws, f'M{i}')
                bFiled_info_Pass = False
            elif field_attr.lower() not in fieldRWOp_arr:
                print("Cell[M"+str(i)+"] must be rw attribute string.")
                markCell_InvalidFunc(ws, f'M{i}')
                bFiled_info_Pass = False

            if bFiled_info_Pass:
                endBit = int(endBit)
                startBit = int(startBit)

                if field_name != 'RESERVED' and reg_info.is_fieldInReg(field_name):
                    print("Field Name if not be \"reserved\" NOT Allowed repeat at Row "+str(i))
                    markCell_InvalidFunc(ws, f'J{i}')
                    bFiled_info_Pass = False

                if endBit is None or startBit is None:
                    print("Field endbit and startbit must be filled at Row "+str(i))
                    bFiled_info_Pass = False
                else:
                    if endBit < startBit:
                        print("Field End Pos must >= Start Pos at Row "+str(i))
                        markCell_InvalidFunc(ws, f'L{i}')
                        bFiled_info_Pass = False
                    if endBit >= laststartBit:
                        print("Field End Pos must < last row Start Pos at Row "+str(i))
                        markCell_InvalidFunc(ws, f'K{i}')
                        bFiled_info_Pass = False
                    laststartBit = startBit

            if bFiled_info_Pass:
                field_inst = St_Filed_info(field_name, field_attr.upper())
                field_inst.xls_row = i
                field_inst.end_bit = endBit
                field_inst.start_bit = startBit
                default_val = row[13]
                field_hdl = row[14]
                field_enum = row[15]
                field_constr = row[16]
                if isinstance(default_val, str):
                    if isHexString(default_val):
                        field_inst.defaultValue = int(default_val, 16)
                if isinstance(field_enum, str):
                    enum_lst = field_enum.splitlines()
                    b_enum_err = False
                    bFirst = True
                    for em in enum_lst:
                        em_val = em.replace(',', '')
                        em_val = em_val.strip()
                        em_item_value=''
                        if em_val.find('=') != -1:
                            (em_item_name, em_str,
                                em_item_value) = em_val.partition('=')
                            em_item_name = em_item_name.strip()
                            em_item_value = em_item_value.strip().upper()
                        else:
                            b_enum_err = True
                        if b_enum_err:
                            break
                        em_item_int_val = 0
                        if em_item_value.isdecimal():
                            em_item_int_val = int(em_item_value)
                        elif isHexString(em_item_value):
                            em_item_int_val = int(em_item_value, 16)
                        else:
                            b_enum_err = True
                            break
                        if bFirst and (field_inst.defaultValue != em_item_int_val):
                            print("Field default value must be the first enum value at Row "+str(i))
                            markCell_InvalidFunc(ws, f'N{i}')
                            bFiled_info_Pass = False
                        bFirst = False
                    if not b_enum_err:
                        field_inst.field_enumstr = field_enum
                    else:
                        print('Field enum val must be "emName =  (dec or hex value) " at Row '+str(i))
                        markCell_InvalidFunc(ws, f'P{i}')
                        bFiled_info_Pass = False
                if isinstance(field_constr, str):
                    field_inst.field_constr = field_constr

                field_inst.hdl_path = field_hdl

                random_enable = row[17]
                comments = row[18]
                if isinstance(comments, str):
                    field_inst.field_comments = comments

                if isinstance(random_enable, int):
                    field_inst.bRandom_Enable = (random_enable == 1)
                reg_info.add_field(field_inst)

            if not bFiled_info_Pass:
                bCheckPass = False

            bNeedNewReg = False
            if laststartBit == 0:
                bNeedNewReg = True

        i += 1

    if bCheckPass and st_module_list:
        mod_inst = st_module_list[0]
        reg_list = mod_inst.reg_list
        for reg in reg_list:
            for fd in reg.field_list:
                if fd in reg_list:
                    markCell_InvalidFunc(ws, f'J{fd.xls_row}')
                    print("Field Name NOT Allow same as Reg Name at row " + str(fd.xls_row))
                    bCheckPass = False
                    break
    if sys.platform== 'win32' and bCheckPass:
        print("Check Sheet:  result Pass")
    return modName, st_module_list, bCheckPass


def output_SequenceSv_moduleFile(module_lst ,modName):
    modName = modName.lower()
    out_sv_module_Name = f'{modName}_v_reg_test_sequence'
    out_sv_file_name = './'+out_sv_module_Name+'.sv'
    if sys.platform == 'linux':
        out_sv_file_name = os.path.join(get_output_dut_cfg_dir(), out_sv_module_Name+'.svh')
    if module_lst:
        module_inst = module_lst[0]
        with open(out_sv_file_name, 'w+') as sv_file:
            fileStr = '//created by xlsFlowX \n\n'
            fileStr += f'class {modName}_v_reg_test_sequence extends cip_base_sequence;\n\n'
            fileStr += f'\t`uvm_object_utils({modName}_v_reg_test_sequence)\n\n'
            fileStr += f'\tfunction new(string name="{modName}_v_reg_test_sequence");\n'
            fileStr += '\t\tsuper.new(name);\n\tendfunction\n\n'
            fileStr += '\tvirtual task body();\n\n'
            fileStr += '\t\tuvm_reg_hw_reset_seq     reg_rst_seq;\n'
            bAllRegFdHdlPathEmpty = True
            modName_U = modName.upper()
            reg_fd_access_str =''
            modinstCount= len(module_lst)
            if isinstance(module_inst,St_Module_info):
                for index in range(modinstCount):
                    fd_hdl_empty_str ='' 
                    module_index = f'{modName_U}{index}'
                    for reg in module_inst.reg_list:
                        if reg.bVirtual:
                            continue
                        bRegFdHdlEmpty = True
                        for fd in reg.field_list:
                            if fd.hdl_path:
                                bAllRegFdHdlPathEmpty = False
                                bRegFdHdlEmpty = False
                        if bRegFdHdlEmpty:
                            fd_hdl_empty_str += '\t\tuvm_resource_db#(bit)::set({"REG::",p_sequencer.u_soc_reg_model.'
                            fd_hdl_empty_str += module_index
                            if reg.bGroup_start and reg.group_size and reg.group_dim:
                                if reg.bGroup_stop:
                                    fd_hdl_empty_str += f'.{reg.reg_name}[{reg.group_dim}]'
                                else:
                                    fd_hdl_empty_str += f'{reg.group_name}[{reg.group_dim}].{reg.reg_name}'
                                    pass
                            else:
                                fd_hdl_empty_str += f'.{reg.reg_name}'
                            fd_hdl_empty_str += '.get_full_name()},"NO_REG_ACCESS_TEST",1,this);\n'
                    if bAllRegFdHdlPathEmpty:
                        pass
                    else:
                        reg_fd_access_str += fd_hdl_empty_str + '\n'
                pass
            if not bAllRegFdHdlPathEmpty:
                fileStr += '\t\tuvm_reg_access_seq       reg_access_seq;\n\n'
            fileStr += '\t\tsuper.body;\n\n'


            fileStr += '\t\t`uvm_info("UVM_SEQ","register reset sequence started",UVM_LOW)\n'
            fileStr += '\t\treg_rst_seq = new();\n'
            for index in range(modinstCount):
                module_index = f'{modName_U}{index}'
                fileStr += f'\t\treg_rst_seq.model = p_sequencer.u_soc_reg_model.{module_index};\n'
                fileStr += '\t\treg_rst_seq.start(p_sequencer);\n'
            fileStr += '\t\t`uvm_info("UVM_SEQ","register reset sequence finished",UVM_LOW)\n\n'

            if not bAllRegFdHdlPathEmpty:
                fileStr += '\t\t`uvm_info("UVM_SEQ","register access sequence started",UVM_LOW)\n'
                fileStr += reg_fd_access_str
                fileStr += '\t\treg_access_seq = new();\n'
                for index in range(modinstCount):
                    module_index = f'{modName_U}{index}'
                    fileStr += f'\t\treg_access_seq.model = p_sequencer.u_soc_reg_model.{module_index};\n'
                    fileStr += '\t\treg_access_seq.start(p_sequencer);\n'
                    pass

                fileStr += '\t\t`uvm_info("UVM_SEQ","register access sequence finished",UVM_LOW)\n'
            
            fileStr += '\n\tendtask: body\n\n'
            fileStr += f'endclass:{modName}_v_reg_test_sequence\n'
            sv_file.write(fileStr)
            pass
    pass

def output_SV_moduleFile(module_inst, modName):
    out_svh_module_Name = modName.lower()+'_dut_cfg'
    out_svh_file_name = './'+out_svh_module_Name+'.svh'
    if sys.platform == 'linux':
        out_svh_file_name = os.path.join(get_output_dut_cfg_dir(), out_svh_module_Name+'.svh')
    
    with open(out_svh_file_name, 'w+') as sv_file:
        heder_str = f'_{modName.upper()}_DUT_CFG_SVH_'
        file_str = F'`ifndef {heder_str}\n`define {heder_str}\n\n'

        file_enum_str = ''

        uvm_field_str = f'\n\t`uvm_object_utils_begin({out_svh_module_Name})\n'
        uvm_fd_val_def_str = ""
        val_def_strarr = ["// Autor: Auto generate by sv", "// Version: 0.0.2 X",
                          "// Description : set module reg field random value", "// Waring: Do NOT Modify it !", "#pragma once"]
        for strarr in val_def_strarr:
            uvm_fd_val_def_str += f'\t\t$fdisplay(fd, "{strarr}" );\n'

        uvm_fd_val_def_str += '\t\t$fdisplay(fd, "   " );\n\n'
        file_cls_str = f'class {out_svh_module_Name} extends uvm_object;\n\n'
        for reg in module_inst.reg_list:
            for fd in reg.field_list:
                if fd.bRandom_Enable and fd.attribute.find('W') != -1:
                    reg_fd_name = f'{reg.reg_name}___{fd.field_name}'
                    b_fd_enum = False
                    nbit_Wid = fd.end_bit-fd.start_bit+1
                    bit_str = 'bit'
                    if nbit_Wid > 1:
                        bit_str = f'bit [{nbit_Wid-1}:0]'
                    if fd.field_enumstr:
                        # print(fd.field_enumstr)
                        b_fd_enum = True
                        enum_lst = fd.field_enumstr.splitlines()
                        file_enum_str += f'typedef enum {bit_str}'+' {\n'
                        b_emFirstitem = True
                        for em in enum_lst:
                            # print(em)
                            em_val = em.replace(',', '')
                            em_val = em_val.strip()
                            (em_item_name, emstr,
                             em_item_value) = em_val.partition('=')
                            em_item_name = em_item_name.strip()
                            em_item_value = em_item_value.strip().upper()
                            if not b_emFirstitem:
                                file_enum_str += ',\n'
                            if len(em_item_value) and em_item_value.startswith('0X'):
                                em_item_value_int = int(em_item_value, 16)
                                file_enum_str += f'\t{em_item_name} {emstr} {em_item_value_int}'
                            else:
                                file_enum_str += f'\t{em_item_name} {emstr} {em_item_value}'
                            b_emFirstitem = False
                            # file_str
                        file_enum_str += '\n} '+f'em_{reg_fd_name};\n\n'

                    if b_fd_enum:
                        file_cls_str += f'\trand em_{reg_fd_name} {reg_fd_name};\n'
                    else:
                        file_cls_str += f'\trand {bit_str}  {reg_fd_name};\n'

                    uvm_field_str += f'\t\t`uvm_field_int({reg_fd_name}, UVM_ALL_ON)\n'

                    fd_name_VAL = f'{reg_fd_name.upper()}_VALUE_'
                    fd_name_VAL = fd_name_VAL.ljust(48)
                    if b_fd_enum:
                        uvm_fd_val_def_str += f'\t\t$fdisplay(fd, "#define \t {fd_name_VAL}   0x%X   //%s",  {reg_fd_name}, {reg_fd_name}.name());\n'
                    else:
                        uvm_fd_val_def_str += f'\t\t$fdisplay(fd, "#define \t {fd_name_VAL}   0x%X",  {reg_fd_name});\n'
        uvm_field_str += f'\t`uvm_object_utils_end\n'

        uvm_field_str += f'\n\tfunction new(string name = "{out_svh_module_Name}");\n'
        uvm_field_str += """\t\tsuper.new(name);
    endfunction:new

    virtual function void print_cfg_to_file();
        int fd;
"""
        uvm_field_str += f'\t\tfd = $fopen("{modName}_dut_cfg.h");\n'

        uvm_field_str += uvm_fd_val_def_str
        uvm_field_str += """
        $fclose(fd);
    endfunction:print_cfg_to_file
"""
        file_cls_str += uvm_field_str
        file_cls_str += """endclass

`endif
"""
        file_str += file_enum_str+'\n\n'
        file_str += file_cls_str
        sv_file.write(file_str)
        sv_file.close()

        return out_svh_file_name


def output_C_moduleFile(st_module_list, module_inst, modName):
    out_C_file_Name = modName+'_reg'
    out_file_name = './'+out_C_file_Name+'.h'
    if sys.platform == 'linux':
        out_file_name = os.path.join(get_output_c_dir(),out_C_file_Name+'.h')

    with open(out_file_name, 'w+') as out_file:
        fileHeader = """// Autor: Auto generate by python From module excel\n
// Version: 0.0.2 X
// Description : struct define for module \n
// Waring: Do NOT Modify it !
#pragma once
"""
        fileHeader += f'#ifndef  _CIP_MODULE_{modName}_DEFINE_\n'
        fileHeader += f'#define  _CIP_MODULE_{modName}_DEFINE_\n\n'
        fileHeader += """
#include <stdint.h>   //for use uint32_t type
#ifdef __cplusplus
extern "C" {
#endif
 
"""

        file_body_str = """#pragma pack(4)
typedef struct {
"""
    #     st_filed_info* pfield;
    # } st_module_{modName};"""
        # 定义module的结构体
        last_offset = 0
        nRegReservedIndex = 0
        field_define_str = """#ifdef _CIP_REG_POS_OPRATION_
////////////////////////////////////////////////////////////////////////
////////////////////define for field opration///////////////////////////
"""
        bRegGroup = False
        group_dim = 0
        group_size = 0
        group_startPos = 0
        group_name = ''
        nRegData_size = module_inst.data_width/8
        uint_str = 'uint32_t'
        # match nRegData_size:
        #     case 1:
        #         uint_str = 'uint8_t'
        #     case 2:
        #         uint_str = 'uint16_t'
        #     case 4:
        #         uint_str = 'uint32_t'
        #     case 8:
        #         uint_str = 'uint64_t'
        if nRegData_size == 1:
            uint_str = 'uint8_t'
        elif nRegData_size == 2:
            uint_str = 'uint16_t'
        elif nRegData_size == 4:
            uint_str = 'uint32_t'
        elif nRegData_size == 8:
            uint_str = 'uint64_t'

        # print('module data_width: {0}'.format(module_inst.data_width))
        group_index = -1
        for reg in module_inst.reg_list:
            if reg.bVirtual:
                continue
            reg_offset = reg.offset
            if reg_offset != last_offset:
                # 增加占位
                nRerived = (reg_offset-last_offset) / nRegData_size
                n = 0
                while n < nRerived:
                    file_body_str += f'\tvolatile {uint_str} u_reg_RESERVED{nRegReservedIndex};  \n'
                    nRegReservedIndex += 1
                    n += 1
            if reg.bGroup_start and reg.group_size and reg.group_dim:
                bRegGroup = True
                group_index = 0
                group_startPos = reg_offset
                group_size = reg.group_size
                group_dim = reg.group_dim
                group_name += reg.reg_name
                file_body_str += "\tvolatile struct  {\n"
            last_offset = reg_offset + nRegData_size
            reg_desc_str = ''
            if reg.desc:
                reg_desc_str = f'/* {reg.desc} */'

            # print('last_offset: is {0}'.format(last_offset))
            field_count = reg.field_count()
            if field_count:
                if bRegGroup:
                    file_body_str += '\t'
                file_body_str += "\tvolatile struct  {\n"
                nFieldReservedIndex = 0
                # filed_name=''
                field_index = field_count-1
                field_bitPos = 0
                while field_index != -1:
                    fd = reg.field_list[field_index]
                    if fd.start_bit != field_bitPos:
                        # 需要补齐field
                        if bRegGroup:
                            file_body_str += '\t'
                        file_body_str += f'\t\t{uint_str} fd_RESERVED{nFieldReservedIndex} : {fd.start_bit-field_bitPos} ;\n'
                        nFieldReservedIndex += 1
                    bReserved = False
                    field_bitPos = fd.end_bit+1
                    fd.field_comments = fd.field_comments.replace(
                        '\n', ' ').replace('\r', ' ')
                    nBitWid = field_bitPos-fd.start_bit
                    filed_name = fd.field_name
                    if filed_name == 'RESERVED':   #'reserved'
                        filed_name = f'RESERVED{nFieldReservedIndex}'
                        nFieldReservedIndex += 1
                        bReserved = True
                    if bRegGroup:
                        file_body_str += '\t'
                    fd_comments_str = ''
                    if fd.field_comments:
                        fd_comments_str = f'/*{fd.field_comments} */'
                    if field_index != 0 or not bReserved:
                        file_body_str += f'\t\t{uint_str} fd_{filed_name} : {nBitWid} ; {fd_comments_str}\n'
                    field_index -= 1
                    if not bReserved:
                        field_str_ = f'{reg.reg_name.upper()}_{filed_name.upper()}'
                        field_define_str += f'//define for {field_str_}\n'
                        field_define_str += f'#define \t {field_str_}_POS \t      {fd.start_bit}U\n'
                        strfdMask = f'{bitWidMask_arr[nBitWid-1]}'
                        field_define_str += f'#define \t {field_str_}_MSK \t      (({uint_str}){strfdMask} << {field_str_}_POS)\n'
                        if fd.attribute.find('W') != -1:
                            field_define_str += f'#define \t {field_str_}_SET(val) \t  (({uint_str})((val) & {strfdMask}) << {field_str_}_POS)\n'

                        field_define_str += f'#define \t {field_str_}_GET(val) \t  (({uint_str})((val) & {field_str_}_MSK) >> {field_str_}_POS)\n'
                        field_define_str += '\n\n'
    # define QSPI_FCMDCR_NMDMYC_POS          7U
    # define QSPI_FCMDCR_NMDMYC_MSK          ((uint32_t)0x1F << QSPI_FCMDCR_NMDMYC_POS)
    # define QSPI_FCMDCR_NMDMYC              QSPI_FCMDCR_NMDMYC_MSK
    # define QSPI_FCMDCR_NMDMYC_SET(val)     ((uint32_t)((val) & 0x1F) << QSPI_FCMDCR_NMDMYC_POS)
    # define QSPI_FCMDCR_NMDMYC_GET(val)     ((uint32_t)((val) & QSPI_FCMDCR_NMDMYC_MSK) >> QSPI_FCMDCR_NMDMYC_POS)

                if bRegGroup:
                    file_body_str += '\t'
                    reg.group_index = group_index
                file_body_str += "\t}\t" + \
                    f'st_reg_{reg.reg_name};   {reg_desc_str} \n'
            else:
                if bRegGroup:
                    file_body_str += '\t'
                file_body_str += f'\tvolatile {uint_str} u_reg_{reg.regname};   {reg_desc_str} */\n'

            if bRegGroup and reg.bGroup_stop:
                if not reg.bGroup_start:
                    group_name += '__'+reg.reg_name
                    # 需要修改该group的其他reg的groupName
                nRerived = (group_size-group_startPos) / nRegData_size
                n = 0
                while n < nRerived:
                    file_body_str += f'\tvolatile {uint_str} u_reg_reserved{nRegReservedIndex}; \n'
                    nRegReservedIndex += 1
                    n += 1
                file_body_str += "\t}\t" + \
                    f'st_group_{group_name} [{group_dim}];   /* group */\n'
                last_offset = group_size*group_dim + group_startPos
                bRegGroup = False
            if bRegGroup:
                group_index += 1

        file_body_str += "}"
        file_body_str += f'st_module_info_{modName};\n'
        file_body_str += "#pragma pack()\n\n"

        # file_body_str += reg_str
        # file_body_str += field_str

        field_define_str += """
////////////////////define for field opration///////////////////////////
#endif // _CIP_REG_POS_OPRATION_
"""
        file_body_str += '\n\n'+field_define_str

        file_body_str += f'\n\n\n#define \t GET_{modName.upper()}_HANDLE   ( (st_module_info_{modName} *) base_addr )\n\n'

        inst_str = """
////////////////////define for module instance///////////////////////////
"""
        nbusAddrindex = 0
        for mo in st_module_list:
            inst_ = f'{modName}_{busTypestr_arr[mo.bus_type]}_baseAddr{nbusAddrindex}'
            file_body_str += f'#define \t {inst_}  \t{hex(mo.bus_baseAddr)}\n'
            inst_str += f'#define \t {modName.upper()}_{nbusAddrindex}    ( (st_module_info_{modName} *) {inst_} )\n'
            nbusAddrindex += 1

        inst_str += """
////////////////////end of define for module instance///////////////////////////
"""
        file_body_str += inst_str

        file_body_str += """
#ifdef __cplusplus
}  //endof extern "C"
#endif
"""

        file_body_str += f'\n#endif //endof  _CIP_MODULE_{modName}_DEFINE_\n'

        out_file.write(fileHeader)
        out_file.write(file_body_str)
        out_file.close()
        return out_file_name


def output_ralf_moduleFile(module_inst, modName):
    out_ralf_file_Name = './' + modName+'.ralf'
    if sys.platform == 'linux':
        out_ralf_file_Name = os.path.join(get_output_ral_dir(), modName+'.ralf')
    with open(out_ralf_file_Name, 'w+') as out_file:
        fileHeader = """# Autor: Auto generate by python From module excel\n
# Version: 0.0.2 X
# Description : struct define for module \n
# Waring: Do NOT Modify it !

"""

        nRegData_size = int(module_inst.data_width/8)
        file_body_str = f'block {modName} ' + \
            ' {\n\tbytes '+f'{nRegData_size};\n'

        # 定义module的结构体
        last_offset = 0
        nRegReservedIndex = 0

        bRegGroup = False
        group_dim = 0
        group_size = 0
        group_startPos = 0
        group_name = ''

        # print('module data_width: {0}'.format(module_inst.data_width))
        group_index = -1
        for reg in module_inst.reg_list:
            # if reg.bVirtual:
            #     continue
            reg_offset = reg.offset
            # if reg_offset != last_offset:
            #     # 增加占位
            #     nRerived = (reg_offset-last_offset) / nRegData_size
            #     n = 0
            #     while n < nRerived:
            #         if bRegGroup:
            #             file_body_str += f'\tregister reg_reserved{nRegReservedIndex}' + \
            #                 ' {\n\t\tbytes '+f'{nRegData_size};\n'+'\t}\n'
            #         nRegReservedIndex += 1
            #         n += 1
            if reg.bGroup_start and reg.group_size and reg.group_dim:
                bRegGroup = True
                group_index = 0
                group_startPos = reg_offset
                group_size = reg.group_size
                group_dim = reg.group_dim
                # group_name += reg.reg_name
                if reg.bGroup_stop:
                    file_body_str += f'\tregister {reg.reg_name}[{group_dim}] @{reg.offset} + {group_size}'+' {\n'
                    bRegGroup = False
                else:
                    file_body_str += f'\tregfile {reg.group_name}[{group_dim}] @{reg.offset} + {group_size}'+' {\n'
                    file_body_str += f'\t\tregister {reg.reg_name} @{reg.offset-group_startPos}'+' {\n' + \
                        '\t\tbytes '+f'{nRegData_size};\n'
            str_Tab = ''
            if bRegGroup:
                str_Tab = '\t'
                if reg.group_index != 0:
                    file_body_str += f'\t\tregister {reg.reg_name} @{reg.offset-group_startPos}'+' {\n' + \
                        ' \t\tbytes '+f'{nRegData_size};\n'
            else:
                if reg.group_index != 0:
                    if reg.bVirtual:
                        file_body_str += f'\tvirtual register  {reg.reg_name} '+' {\n'
                    else:
                        file_body_str += f'\tregister  {reg.reg_name} @{reg.offset}'+' {\n'
                    file_body_str += F'\t\tbytes {nRegData_size};\n'

            # last_offset = reg_offset + nRegData_size

            # print('last_offset: is {0}'.format(last_offset))
            field_count = reg.field_count()
            if field_count:
                # if bRegGroup:
                #     file_body_str += f'\t\tregister {reg.reg_name} @{reg.offset-group_startPos}'+' {\n' + \
                #     ' \t\tbytes '+f'{nRegData_size};\n'
                # else:
                #     file_body_str += f'\tregister {reg.reg_name} @{reg.offset}'+' {\n' + \
                #     ' \t\tbytes '+f'{nRegData_size};\n'
                # nFieldReservedIndex = 0

                field_index = field_count-1
                # field_bitPos = 0
                while field_index != -1:
                    fd = reg.field_list[field_index]
                    # if fd.start_bit != field_bitPos:
                    #     # 需要补齐field
                    #     if bRegGroup:
                    #         file_body_str += '\t'
                    #     file_body_str += f'\t\tfield fd_reserved{nFieldReservedIndex} '+' {\n\t\t\t'
                    #     if bRegGroup:
                    #         file_body_str += '\t'
                    #     file_body_str += f'bits: {fd.start_bit-field_bitPos} ;\n'
                    #     if bRegGroup:
                    #         file_body_str += '\t'
                    #     file_body_str += '\t\t}'
                    #     nFieldReservedIndex += 1
                    bReserved = False
                    field_bitPos = fd.end_bit+1
                    fd.field_comments = fd.field_comments.replace(
                        '\n', ' ').replace('\r', ' ')
                    nBitWid = field_bitPos-fd.start_bit
                    if fd.field_name == 'RESERVED':   #reserved
                        # fd.field_name = f'reserved{nFieldReservedIndex}'
                        # nFieldReservedIndex += 1
                        bReserved = True
                    if not bReserved:
                        if fd.hdl_path:
                            file_body_str += f'{str_Tab}\t\tfield fd_{fd.field_name} ({fd.hdl_path})  @{fd.start_bit}'+' {\n'
                        else:
                            file_body_str += f'{str_Tab}\t\tfield fd_{fd.field_name} @{fd.start_bit}'+' {\n'
                        file_body_str += f'{str_Tab}\t\t\tbits {nBitWid} ;\n'
                        file_body_str += f'{str_Tab}\t\t\treset {fd.defaultValue} ;\n'
                        if fd.attribute:
                            file_body_str += f'{str_Tab}\t\t\taccess {fd.attribute.lower()} ;\n'
                        if fd.field_enumstr:
                            # print(fd.field_enumstr)
                            # b_fd_enum = True
                            enum_lst = fd.field_enumstr.splitlines()
                            field_enum_str = f'{str_Tab}\t\t\tenum '+' {\n'
                            b_emFirstitem = True
                            for em in enum_lst:
                                # print(em)
                                em_val = em.replace(',', '')
                                em_val = em_val.strip()
                                (em_item_name, emstr,
                                 em_item_value) = em_val.partition('=')
                                em_item_name = em_item_name.strip()
                                em_item_value = em_item_value.strip().upper()
                                if not b_emFirstitem:
                                    field_enum_str += ',\n'
                                if em_item_value and em_item_value.startswith('0X'):
                                    em_item_value_int = int(em_item_value, 16)
                                    field_enum_str += f'{str_Tab}\t\t\t\t{em_item_name} {emstr} {em_item_value_int}'
                                else:
                                    field_enum_str += f'{str_Tab}\t\t\t\t{em_item_name} {emstr} {em_item_value}'
                                b_emFirstitem = False
                                # file_str
                            field_enum_str += f'\n{str_Tab}\t\t\t'+'} \n'
                            file_body_str += field_enum_str
                        file_body_str += f'{str_Tab}\t\t\tconstraint c_st_{fd.field_name} ' + '{\n\t\t\t'
                        file_body_str += str_Tab+'}\n'

                        file_body_str += str_Tab+'\t\t}'
                        if fd.field_comments:
                            file_body_str += f'; #{fd.field_comments} \n'
                        else:
                            file_body_str += '\n'
                    field_index -= 1
                # if bRegGroup:
                #     reg.group_index = group_index
                file_body_str += str_Tab+'\t}; ' + f' #{reg.desc} \n'
            else:
                file_body_str += f'{str_Tab}\tregister  {reg.regname} @{reg.offset-group_startPos}'+' {\n'
                file_body_str += F'{str_Tab}\t\tbytes {nRegData_size};\n'
                file_body_str += str_Tab+'\t\tfield reserved {\n'
                file_body_str += str_Tab+'\t\t\tbits 4 ;\n'
                file_body_str += str_Tab+'\t\t}\n'
                file_body_str += str_Tab+'\t}'+f';  #{reg.desc}\n'

            if bRegGroup and reg.bGroup_stop:
                # if not reg.bGroup_start:
                #     group_name += '__'+reg.reg_name
                #     # 需要修改该group的其他reg的groupName
                # nRerived = (group_size-group_startPos) / nRegData_size
                # n = 0
                # while n < nRerived:
                #     file_body_str += f'\tregister  reg_reserved{nRegReservedIndex} ' + '{\n'
                #     file_body_str += F'\t\tbytes {nRegData_size};\n'
                #     file_body_str += '\t}'+f';  #{reg.desc}\n'
                #     nRegReservedIndex += 1
                #     n += 1
                file_body_str += "\t} ; " + \
                    f'#group_{group_name} [{group_dim}]\n'
                # last_offset = group_size*group_dim + group_startPos
                bRegGroup = False
            if bRegGroup:
                group_index += 1

        file_body_str += '\n}; # End of block module '+modName+'\n'
        out_file.write(fileHeader)
        out_file.write(file_body_str)
        out_file.close()
        return out_ralf_file_Name


def outModuleFieldDefaultValueCheckCSrc(module_inst_list, modName):
    # print(modName)
    out_C_file_Name=''
    if sys.platform == 'win32':
        dirName = './module_check_defaultvalue/'+modName
        if not os.path.exists(dirName):
            os.makedirs(dirName)
        out_C_file_Name = dirName+'/main.c'
    elif sys.platform == 'linux':
        dirName = os.path.join(get_output_c_dir(), modName.lower()+'_reg_c_reg_test')
        out_C_file_Name = dirName+'_main.c'
    with open(out_C_file_Name, 'w+') as out_file:
        fileHeader = """// Autor: Auto generate by python From module excel\n
// Version: 0.0.2 X
// Description : field default value check for module instance \n
// Waring: Do NOT Modify it !
// Copyright (C) 2020-2021 CIP United Co. Ltd.  All Rights Reserved.

#define DEBUG
//#define INFO
#define WARNING
#define NOTICE
#define ERROR
#define PASS
#define FAIL

#define CHECK_MODULE_FIELD_DEFAULT_VALUE

#define CHECK_MOUDLE_FIELD_WRITE_VALUE

#include "log.h"
#include "pll.h"

"""
        filebodystr = f'#include "{modName}_reg.h"\n'
        filebodystr += """
int main()
{
    //printf("enter main.\\n");
    uAptiv_clk_init();
"""
        mod_inst_name = modName.upper()
        mod_inst = module_inst_list[0]
        filebodystr += f'\tprintf("After clock switch, Now Check Module: {modName}.\\n");\n'
        if mod_inst.data_width > 32:
            filebodystr += f'\tuint64_t nRegFdVal = 0;\n'
        else:
            filebodystr += f'\tunsigned int nRegFdVal = 0;\n'
        mod_count = len(module_inst_list)

        filebodystr += f'\n\tunsigned int nTotalErr = 0;\n'
        if mod_count > 1:
            filebodystr += f'\tst_module_info_{modName} * module_inst[{mod_count}] = ' + \
                '{'+f'{mod_inst_name}_0'
            for i in range(1, mod_count):
                filebodystr += f'\n\t\t,{mod_inst_name}_{i}'
            filebodystr += '};\n\n'
            filebodystr += '#ifdef CHECK_MODULE_FIELD_DEFAULT_VALUE\n'
            filebodystr += f'\tunsigned int nErrCount_default[{mod_count}] = '+'{0};\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

            filebodystr += '#ifdef CHECK_MOUDLE_FIELD_WRITE_VALUE\n'
            filebodystr += f'\tunsigned int nErrCount_wirte[{mod_count}] = '+'{0};\n'
            filebodystr += '#endif // CHECK_MOUDLE_FIELD_WRITE_VALUE\n\n'

            filebodystr += f'\tfor(int i = 0; i < {mod_count}; ++i)\n'
            filebodystr += '\t{\n'
            filebodystr += '#ifdef CHECK_MODULE_FIELD_DEFAULT_VALUE\n'
            modinst_var = 'module_inst[i]'
            errCount_var = 'nErrCount_default[i]'
            errCount_write_var = 'nErrCount_wirte[i]'
            str1, str2 = getModuleFdStr(
                mod_inst, errCount_var, errCount_write_var, modinst_var)
            filebodystr += str1
            filebodystr += '\t\tif(nErrCount_default[i])\n'
            filebodystr += '\t\t{\n'
            filebodystr += f'\t\t\tError("Inst_%u def-Vals have [%u] fds NOT Same!\\n", i, nErrCount_default[i]);\n'
            filebodystr += '\t\t\tnTotalErr += nErrCount_default[i];\n'
            filebodystr += '\t\t}\n'
            filebodystr += f'\t\telse\n\t\t\tNotice("Inst_%u def-Vals are OK!\\n", i);\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

            filebodystr += str2
            filebodystr += '\t\tif(nErrCount_wirte[i])\n'
            filebodystr += '\t\t{\n'
            filebodystr += f'\t\t\tError("Inst_%u write-Vals have [%u] fds NOT Same!\\n", i, nErrCount_wirte[i]);\n'
            filebodystr += '\t\t\tnTotalErr += nErrCount_wirte[i];\n'
            filebodystr += '\t\t}\n'
            filebodystr += f'\t\telse\n\t\t\tNotice("Inst_%u write-Vals are OK!\\n", i);\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'
            filebodystr += '\t}\n'
        elif mod_count == 1:
            filebodystr += f'\tst_module_info_{modName} * module_inst = {mod_inst_name}_0 ;\n'
            filebodystr += '#ifdef CHECK_MODULE_FIELD_DEFAULT_VALUE\n'
            filebodystr += f'\tunsigned int nErrCount_default = 0;\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

            filebodystr += '#ifdef CHECK_MOUDLE_FIELD_WRITE_VALUE\n'
            filebodystr += f'\tunsigned int nErrCount_wirte = 0;\n'
            filebodystr += '#endif // CHECK_MOUDLE_FIELD_WRITE_VALUE\n\n'

            modinst_var = 'module_inst'
            errCount_var = 'nErrCount_default'
            errCount_write_var = 'nErrCount_wirte'
            # filebodystr += getModuleFdStr(mod_inst,
            #                               errCount_var, modinst_var, False)
            str1, str2 = getModuleFdStr(
                mod_inst, errCount_var, errCount_write_var, modinst_var, False)

            filebodystr += str1

            filebodystr += '\t\tif(nErrCount_default)\n'
            filebodystr += '\t\t{\n'
            filebodystr += f'\t\t\tError("Inst_%u def-Vals have [%u] fds NOT Same!\\n", i, nErrCount_default);\n'
            filebodystr += '\t\t\tnTotalErr += nErrCount_default;\n'
            filebodystr += '\t\t}\n'
            filebodystr += f'\t\telse\n\t\t\tNotice("Inst_%u def-Vals are OK!\\n", i);\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

            filebodystr += str2
            filebodystr += '\t\tif(nErrCount_wirte)\n'
            filebodystr += '\t\t{\n'
            filebodystr += f'\t\t\tError("Inst_%u write-Vals have [%u] fds NOT Same!\\n", i, nErrCount_wirte);\n'
            filebodystr += '\t\t\tnTotalErr += nErrCount_wirte;\n'
            filebodystr += '\t\t}\n'
            filebodystr += f'\t\telse\n\t\t\tNotice("Inst_%u write-Vals are OK!\\n", i);\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

       
        filebodystr += f'\tif(nTotalErr == 0)\n'
        filebodystr += f'\t\tPass("{modName} Vals OK!\\n");\n'
        filebodystr += f'\telse\n\t\tFail("{modName} Vals Not OK!\\n");\n'
        filebodystr += '\n\treturn 0;\n}\n'
        out_file.write(fileHeader)
        out_file.write(filebodystr)
        out_file.close()

        return out_C_file_Name


def getModuleFdStr(mod_inst, errCount_var, errCount_Write_var, modinst_var, bForLoop=True):
    filebodystr = ''
    fieldWriteCheckstr = '#ifdef CHECK_MOUDLE_FIELD_WRITE_VALUE\n'
    group_dim = 0
    str_Tab = ''
    if bForLoop:
        str_Tab = '\t'
    # mod_name = mod_inst.module_name
    for reg in mod_inst.reg_list:
        if reg.bVirtual:
            continue
        if reg.bGroup_start and reg.group_dim:
            group_dim = reg.group_dim

        if reg.group_index >= 0 and reg.group_name:
            for g_i in range(0, group_dim):
                for fd in reg.field_list:
                    fd_name=fd.field_name.upper()
                    if fd_name.startswith('RESERVED'): #reserved
                        continue

                    reg_fd_var = f'{reg.reg_name}.fd_{fd.field_name}'
                    group_name = reg.group_name[8:]
                    fd_var = f'gp{group_name}[{g_i}].{reg_fd_var}'
                    module_fd_var = f'{str_Tab}{modinst_var}->{reg.group_name}[{g_i}].st_reg_{reg_fd_var}'
                    nBitWid = fd.end_bit-fd.start_bit+1
                    if fd.attribute.find('R') != -1:
                        filebodystr += f'{str_Tab}\tnRegFdVal = {module_fd_var};\n'
                        filebodystr += f'{str_Tab}\tif(nRegFdVal != {fd.defaultValue})\n{str_Tab}'
                        filebodystr += '\t{\n'

                        if bForLoop:
                            filebodystr += f'{str_Tab}\t\tError("Inst_%u # {fd_var}  [0x%X] is NOt same! \\n", i, nRegFdVal);\n'
                        else:
                            filebodystr += f'{str_Tab}\t\tError("{fd_var}  [0x%X] is NOt same! \\n", nRegFdVal);\n'
                        filebodystr += f'{str_Tab}\t\t++{errCount_var};\n{str_Tab}'
                        filebodystr += '\t}\n'
                        if bForLoop:
                            filebodystr += f'{str_Tab}\telse\n{str_Tab}\t\tInfo("Inst_%u # {fd_var} Value is OK. \\n", i);\n'
                        else:
                            filebodystr += f'{str_Tab}\telse\n{str_Tab}\t\tInfo("{fd_var} Value is OK. \\n");\n'

                    enum_val_lst = []
                    if fd.field_enumstr:
                        # print(fd.field_enumstr)
                        enum_lst = fd.field_enumstr.splitlines()
                        for em in enum_lst:
                            # print(em)
                            em_val = em.replace(',', '')
                            em_val = em_val.strip()
                            (em_item_name, emstr,
                             em_item_value) = em_val.partition('=')
                            # em_item_name = em_item_name.strip()
                            em_item_value = em_item_value.strip().upper()
                            if em_item_value:
                                if em_item_value.startswith('0X'):
                                    em_item_value_int = int(em_item_value, 16)
                                else:
                                    em_item_value_int = int(em_item_value)
                                enum_val_lst.append(em_item_value_int)
                    # 先赋值为全1
                    if fd.attribute.find('W') != -1:
                        if len(enum_val_lst) > 1:
                            strfdMask = enum_val_lst[-1]
                            fieldWriteCheckstr += fieldWriteChk_func(
                                errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)

                            strfdMask = enum_val_lst[0]
                            fieldWriteCheckstr+=fieldWriteChk_func(
                                errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)
                        else:
                            strfdMask = f'{bitWidMask_arr[nBitWid-1]}'
                            fieldWriteCheckstr += fieldWriteChk_func(
                                errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)

                            strfdMask = 0
                            fieldWriteCheckstr += fieldWriteChk_func(
                                errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)

        else:
            for fd in reg.field_list:
                fd_name=fd.field_name.upper()
                if fd_name.startswith('RESERVED'): #reserved
                    continue
                # 根据attr，判断是否可读

                fd_var = f'{reg.reg_name}.fd_{fd.field_name}'
                module_fd_var = f'{modinst_var}->st_reg_{fd_var}'
                nBitWid = fd.end_bit-fd.start_bit+1

                if fd.attribute.find('R') != -1:
                    filebodystr += f'{str_Tab}\tnRegFdVal = {module_fd_var};\n'
                    filebodystr += f'{str_Tab}\tif(nRegFdVal != {fd.defaultValue})\n{str_Tab}'
                    filebodystr += '\t{\n'

                    if bForLoop:
                        filebodystr += f'{str_Tab}\t\tError("Inst_%u # {fd_var}  [0x%X] is not same! \\n", i, nRegFdVal);\n'
                    else:
                        filebodystr += f'{str_Tab}\t\tError("{fd_var}  [0x%X] is not same! \\n", nRegFdVal);\n'
                    filebodystr += f'{str_Tab}\t\t++{errCount_var};\n{str_Tab}'
                    filebodystr += '\t}\n'
                    if bForLoop:
                        filebodystr += f'{str_Tab}\telse\n{str_Tab}\t\tInfo("Inst_%u # {fd_var} Value is OK. \\n", i);\n'
                    else:
                        filebodystr += f'{str_Tab}\telse\n{str_Tab}\t\tInfo("{fd_var} Value is OK. \\n");\n'

                enum_val_lst = []
                if fd.field_enumstr:
                    # print(fd.field_enumstr)
                    enum_lst = fd.field_enumstr.splitlines()
                    for em in enum_lst:
                        # print(em)
                        em_val = em.replace(',', '')
                        em_val = em_val.strip()
                        (em_item_name, str, em_item_value) = em_val.partition('=')
                        # em_item_name = em_item_name.strip()
                        em_item_value = em_item_value.strip().upper()
                        if em_item_value:
                            if em_item_value.startswith('0X'):
                                em_item_value_int = int(em_item_value, 16)
                            else:
                                em_item_value_int = int(em_item_value)
                            enum_val_lst.append(em_item_value_int)

                if fd.attribute.find('W') != -1:
                    if len(enum_val_lst) > 1:
                        strfdMask = enum_val_lst[-1]
                        fieldWriteCheckstr += fieldWriteChk_func(
                            errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)

                        strfdMask = enum_val_lst[0]
                        fieldWriteCheckstr += fieldWriteChk_func(
                            errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)
                    else:
                        strfdMask = f'{bitWidMask_arr[nBitWid-1]}'

                        fieldWriteCheckstr += fieldWriteChk_func(
                            errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)

                        strfdMask = 0
                        fieldWriteCheckstr += fieldWriteChk_func(
                            errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)

        if reg.bGroup_stop:
            group_dim = 0

    # fieldWriteCheckstr += '#endif //CHECK_MOUDLE_FIELD_WRITE_VALUE\n'
    return filebodystr, fieldWriteCheckstr


def fieldWriteChk_func(errCount_Write_var, str_Tab, fd_var, module_fd_var, strfdMask):
    fdWriteCheckstr = ''
    fdWriteCheckstr += f'{str_Tab}\t{module_fd_var} = {strfdMask};\n'
    fdWriteCheckstr += f'{str_Tab}\tnRegFdVal = {module_fd_var};\n'
    fdWriteCheckstr += f'{str_Tab}\tif({module_fd_var} != {strfdMask})\n{str_Tab}'
    fdWriteCheckstr += '\t{\n'
    fdWriteCheckstr += f'{str_Tab}\t\tError("Inst_%u # {fd_var}  [0x%X] NOt same as Write [{strfdMask}]! \\n", i, nRegFdVal);\n'
    fdWriteCheckstr += f'{str_Tab}\t\t++{errCount_Write_var};\n{str_Tab}'
    fdWriteCheckstr += '\t}\n'
    return fdWriteCheckstr


def dealwith_excel(xls_file):
    # "UART_final_202301010.xls"
    wb = load_workbook(xls_file)
    ws = wb.active
    modName, st_module_list, bCheckPass = checkModuleSheetVale(ws)
    if bCheckPass:
        mod_inst_count = len(st_module_list)
        if mod_inst_count:
            module_inst = st_module_list[0]
            print('module name: {0}.'.format(modName))

            out_file_list = []
            out_file_name = output_C_moduleFile(
                st_module_list, module_inst, modName)
            if (out_file_name):
                out_file_list.append(out_file_name)

            out_file_name = output_SV_moduleFile(module_inst, modName)
            if (out_file_name):
                out_file_list.append(out_file_name)

            ahb_pos =  len(st_module_list)
            for index in range(mod_inst_count):
                if st_module_list[index].bus_type:
                    ahb_pos = index
                    break
            # print(ahb_pos)
            out_file_name = output_SequenceSv_moduleFile(st_module_list[0:ahb_pos], modName)
            if (out_file_name):
                out_file_list.append(out_file_name)
           
            out_file_name = outModuleFieldDefaultValueCheckCSrc(st_module_list[0:ahb_pos], modName)
            if (out_file_name):
                out_file_list.append(out_file_name)

            out_file_name = output_ralf_moduleFile(module_inst, modName)
            if (out_file_name):
                out_file_list.append(out_file_name)

            # outModuleFieldDefaultValueCheckCSrc(st_module_list[0:1], modName)

            # for module in st_module_list:
            #     print(module.module_info_str())
            # 实例化各个module

            return out_file_list
    else:
        print("Check Failed. Please review the excel file and fix it.")
        filename = os.path.basename(xls_file)
        out_mark_xlsx_file = filename.replace('.xlsx', '_errMk.xlsx')
        # print(out_mark_xlsx_file)
        print("You can review the error mark file {0}.".format(
            out_mark_xlsx_file))
        wb.save(out_mark_xlsx_file)


if __name__ == '__main__':
    # 全路径是为方便在vscode中进行调试
    # file_name = 'D:/workspace/demopy/excel_flow/excel/ahb_cfg_20230925.xlsx'
    file_name = './UART_CFG_XY2_20231127.xlsx'
    dealwith_excel(file_name)
